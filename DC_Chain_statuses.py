import pandas as pd
import numpy as np
from os.path import exists as file_exists
from openpyxl import load_workbook
import sys

in_file = {
    'name': 'Materialmasterdata.xlsx',
    'status_sheet': 'Report',
    'active_products': 'ZP02',
    'pl_plant_sheet': 'PL_Plant',
    'autoreplacement': 'Autoreplacement',
    # 'stat_definitions': 'Stat_definitions',
    'stat_cross_ref': 'Stat_cross_ref'
}

out_file = {
    'name': 'Status_verification.xlsx',
    'status_sheet': 'Statuses', 
    'autoreplacement': 'Autoreplacement'
    }
            
print(f'From {in_file["name"]} take:\n',
      f'statuses - from {in_file["status_sheet"]}\n',
      f'active products - from {in_file["active_products"]}\n',
      f'statuses corelation - from {in_file["stat_cross_ref"]}\n',
      f'autoreplacement schema - from {in_file["autoreplacement"]}\n',
      f'delivering plant for PL - from {in_file["pl_plant_sheet"]}')


def read_in_file(in_file: str, in_SheetName: str='', header_row: int=0):
    print(f'Reading page {in_SheetName}...', end='')

    try:
        wb = pd.read_excel(in_file, sheet_name=None, header=None)
        sheets = list(wb.keys())
        if in_SheetName == '':
            print("Loading last sheet in file")
            df = wb.get(sheets[-1])
        else:
            df = wb.get(in_SheetName)
        del(wb)
        
    except TypeError:
        wb = load_workbook(filename=in_file, data_only=True)
        if in_SheetName == '':
            print("Loading last sheet in file")
            df = pd.DataFrame(wb[wb.sheetnames[-1]].values)

        else:
            df = pd.DataFrame(wb[in_SheetName].values)

        sheets = wb.sheetnames
        wb.close()

    except FileNotFoundError:
        input(f'\nFile {in_file} not found. Press any key to exit.')
        sys.exit(1)
    
    except Exception as e:
        print('\n', e)
        input('Press any key to exit.')
        sys.exit(1)

    column_names = rename_columns(df.iloc[header_row].values)
    df.columns = column_names
    df.drop(index=[i for i in range(header_row+1)], inplace=True)
    df.reset_index(drop=True, inplace=True)
    
    print('Done.')
    return df, sheets


def rename_columns(column_names):
    for item in enumerate(column_names):
        try:
            column_names[item[0]] = '_'.join(item[1].split())
        except:
            column_names[item[0]] = item[1]
    return column_names


df_statuses, sheets = read_in_file(
    in_file['name'], in_file['status_sheet'], header_row=0)
df_products, _ = read_in_file(
    in_file['name'], in_file['active_products'], header_row=0)
df_pl_plant, _ = read_in_file(
    in_file['name'], in_file['pl_plant_sheet'], header_row=0)
df_auroreplacement, _ = read_in_file(
    in_file['name'], in_file['autoreplacement'], header_row=0)
df_stats_ref, _ = read_in_file(
    in_file['name'], in_file['stat_cross_ref'], header_row=0)


def df_preparation_pipeline():
    global df_statuses
    global df_products
    global df_pl_plant
    global df_auroreplacement
    global df_stats_ref

    df_statuses = df_statuses[['Material', 'Plant', 'Product_line', 
                            'X-distr.chain_status', 'Plant-Sp.Matl_Status', 'DChain_Status']]

    df_products = df_products[['Material_Number',
                            'Text_Material', 'Amount_(Currency)']]
    # if several prices with different validity dates exists - take first record
    df_products.drop_duplicates(
        subset='Material_Number', keep='first', inplace=True)
    df_products.set_index('Material_Number', inplace=True)

    df_pl_plant.set_index(df_pl_plant.columns[0], inplace=True)
    df_pl_plant.dropna(how='all', inplace=True)
    df_pl_plant['Plant_accepted'] = df_pl_plant.transpose().apply(
        lambda x: x.dropna().index.to_list())
    df_pl_plant = df_pl_plant[['Plant_accepted']]

    df_auroreplacement = df_auroreplacement[[
        'Material_Entered', 'Danf_Material_number']]

    df_stats_ref.set_index(df_stats_ref.columns[0], inplace=True)
    df_stats_ref['B600_accepted'] = df_stats_ref.transpose().apply(
        lambda x: x.dropna().index.to_list())
    df_stats_ref = df_stats_ref[['B600_accepted']]

def df_merge_pipeline():
    global df_statuses
    global df_products
    global df_pl_plant
    global df_stats_ref
    global df_output

    df_statuses = df_statuses.join(
        df_products, on='Material', how='outer').reset_index(drop=True)
    df_statuses[['X-distr.chain_status', 'Plant-Sp.Matl_Status',  'DChain_Status']
                ] = df_statuses[['X-distr.chain_status', 'Plant-Sp.Matl_Status', 'DChain_Status']].replace(np.nan, 0)
    df_statuses = df_statuses.astype(
        {'X-distr.chain_status': 'int64', 'Plant-Sp.Matl_Status': 'int64', 'DChain_Status': 'int64'})

    df_statuses = df_statuses.join(
        df_stats_ref, on='X-distr.chain_status', how='left')
    df_statuses['B600_accepted'] = df_statuses['B600_accepted'].apply(
        lambda x: x if isinstance(x, list) else [])

    df_statuses = df_statuses.join(df_pl_plant, on='Product_line', how='left')
    df_statuses['Plant_accepted'] = df_statuses['Plant_accepted'].apply(
        lambda x: x if isinstance(x, list) else [])

    # local statuses are the same? (recomended to be the same)
    df_statuses['Is_local_stat_equal'] = df_statuses['Plant-Sp.Matl_Status'] == df_statuses['DChain_Status']

    # DChain statuses are the same?
    df_statuses['Is_X-DC_equal'] = df_statuses['X-distr.chain_status'] == df_statuses['DChain_Status']

    # is current DC_status within accepted for B600 variants?
    df_statuses['Is_Status_accepted'] = df_statuses[['DChain_Status', 'B600_accepted']].apply(
        lambda row: row.iloc[0] in row.iloc[1], axis=1)

    # is current DC_status 'blocked' if no price?
    df_statuses['Is_Status_accepted'] = df_statuses[['Is_Status_accepted', 'DChain_Status', 'Amount_(Currency)']].apply(
        lambda row: row.iloc[0] if row.iloc[2] else row.iloc[1] in [75, 78], axis=1)

    # is current Plant within accepted for B600 variants?
    df_statuses['Is_Plant_correct'] = df_statuses[[
        'Plant', 'Plant_accepted']].apply(lambda row: row.iloc[0] in row.iloc[1], axis=1)

    df_plant_ok = df_statuses[df_statuses['Is_Plant_correct'] == True]
    df_plant_wrong = df_statuses.drop(df_plant_ok.index)

    # from df_plant_wrong remove products that exist in df_plant_ok
    df_plant_wrong = df_plant_wrong.drop(
        df_plant_wrong[df_plant_wrong['Material'].isin(df_plant_ok['Material'].unique())].index)

    # from df_plant_wrong remove products without price
    df_plant_wrong = df_plant_wrong[~df_plant_wrong['Amount_(Currency)'].isna()]

    df_output = pd.concat([df_plant_wrong, df_plant_ok])

    df_output = df_output[df_output[['Amount_(Currency)', 'B600_accepted', 'Plant_accepted']].any(axis=1)]


def autoreplacement_evaluation():
    global df_auroreplacement
    global df_output

    df_statuses = df_output[['Material', 'Product_line', 'X-distr.chain_status', 'DChain_Status', 'Text_Material']]
    df_statuses.set_index('Material', inplace=True)

    df_auroreplacement = df_auroreplacement.join(
        df_statuses, on='Material_Entered', how='left', rsuffix='_Old')

    df_auroreplacement = df_auroreplacement.join(
        df_statuses, on='Danf_Material_number', how='left', rsuffix='_New')
    
    # # order of steps matters as overwrite expected!!!
    df_auroreplacement['ToDo'] = None

    df_auroreplacement['ToDo'] = df_auroreplacement[['DChain_Status', 'DChain_Status_New', 'ToDo']].apply(
        lambda x: 'cancel replacement' if (np.isnan(x['DChain_Status']) or np.isnan(x['DChain_Status_New'])) else x['ToDo'], axis='columns')

    df_auroreplacement['ToDo'] = df_auroreplacement[['DChain_Status', 'DChain_Status_New', 'ToDo']].apply(
        lambda x: 'review old product status' if (x['DChain_Status'] in [57, 78] and x['ToDo'] is None) else x['ToDo'], axis='columns')

    df_auroreplacement['ToDo'] = df_auroreplacement[['DChain_Status', 'DChain_Status_New', 'ToDo']].apply(
        lambda x: 'review replacement' if (x['DChain_Status_New'] in [57, 75, 78] and x['ToDo'] is None) else x['ToDo'], axis='columns')
    
    df_auroreplacement['ToDo'] = df_auroreplacement[['DChain_Status', 'DChain_Status_New', 'ToDo']].apply(
        lambda x: 'OK' if (x['DChain_Status_New'] == 51 and x['ToDo'] is None) else x['ToDo'], axis='columns')
    
    # # print(df_auroreplacement.head())
    # sys.exit(1)

df_preparation_pipeline()
df_merge_pipeline()
autoreplacement_evaluation()

if file_exists(out_file['name']):
    print(f'Replacing {out_file["name"]}\n')
else:
    print(f'Creating {out_file["name"]}\n')

with pd.ExcelWriter(out_file["name"], mode='w') as writer:
    df_output.to_excel(writer, sheet_name=out_file["status_sheet"], index=False)
    df_auroreplacement.to_excel(
        writer, sheet_name=out_file["autoreplacement"], index=False)

input('Done. Press any key to exit...')
